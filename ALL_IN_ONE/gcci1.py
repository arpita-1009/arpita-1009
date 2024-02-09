from openpyxl import Workbook, load_workbook
import time, os , re
from PIL import Image, ImageOps

import pytesseract
path_to_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = path_to_tesseract


def cut_image(image_path):
    # Open the image
    image = Image.open(image_path)

    # Sort the x-coordinates in ascending order
    x1, x2 = 510 , 1050

    # Cut the image into two parts
    image.crop((0, 0, x1, image.height)).save('left_side.jpg')
    image.crop((x1, 0, x2, image.height)).save('right_side.jpg')
    # image.crop((x2, 0, image.width, image.height)).save('part3.jpg')

    # # Return the three parts as PIL Image objects
    # return part1, part2, part3


def extract_company_details(text):
    company_name_match = re.search(r'([A-Z\s.&.-]+)', text)
    company_phone_match = re.search(r'Phone : (\d{3}-\d{8})', text)

    company_name = company_name_match.group().strip() if company_name_match else None
    company_phone = company_phone_match.group(1) if company_phone_match else None

    return company_name, company_phone



def extract_name_and_phone(text):
    person_name_and_number_pattern = r"([A-Za-z]+(?:\.[A-Za-z]+)?\s[A-Za-z]+(?:\s[A-Za-z]+)?)\s\((\d+)\)"

    matches = re.findall(person_name_and_number_pattern, text)

    person_name_list = []
    mobile_number_list = []
    for match in matches:
        name = match[0]
        mobile_number = match[1]
        person_name_list.append(name)
        mobile_number_list.append(mobile_number)

    # return person_name_list , mobile_number_list
    return person_name_list , mobile_number_list
'''---------------------------------------------'''

# def extract_and_save_data_to_excel(filename):
#     # for part1.jpg image
#     image_1 = Image.open('left_side.jpg')

#     # Extract text from the image
#     text_1 = pytesseract.image_to_string(image_1)


#     # Find Details from Extracted Text to Insert in Excel

#     firm_name , firm_contact_number = extract_company_details(text_1)


#     print("Firm Name : " ,firm_name)
#     print("Firm Contact Number : " , firm_contact_number)

#     # 'part1.jpg' extraction complete

#     '''--------------------------------------------------------'''

#     # for part2.jpg image
#     image_2 = Image.open('right_side.jpg')

#     # Extract text from the image
#     text_2 = pytesseract.image_to_string(image_2)

#     # Find Details from Extracted Text to Insert in Excel
#     person_name , mobile_number = extract_name_and_phone(text_2)

#     print(person_name)
#     print(mobile_number)

#     from openpyxl import Workbook, load_workbook

#     # Function to create a new Excel file with headers if it doesn't exist
#     def create_excel_file(filename):
#         workbook = Workbook()
#         sheet = workbook.active
#         headers = ["Firm Name", "Firm Contact Number", "Person Name", "Mobile Number"]
#         sheet.append(headers)
#         workbook.save(filename)

#     # Function to add a new row to the existing Excel file
#     def add_row_to_excel(filename, data):
#         workbook = load_workbook(filename)
#         sheet = workbook.active
#         sheet.append(data)
#         workbook.save(filename)
#     # from app import get_date
#     # path=get_date()
#     # Save the modified DataFrame back to Excel
#     # path_1 = path[0].replace('.pdf','.xlsx')
#     # filename = "test_23_oct.xlsx"

#     # Check if the file exists, if not, create a new file with headers
#     try:
#         workbook = load_workbook(filename)
#     except FileNotFoundError:
#         create_excel_file(filename)

#     # Check if person_name and mobile_number are lists
#     if isinstance(person_name, list) and isinstance(mobile_number, list):
#         # Make sure the lengths of the lists are the same
#         if len(person_name) == len(mobile_number):
#             # Iterate over the lists and create a new row for each element
#             for i in range(len(person_name)):
#                 data = [firm_name, firm_contact_number, person_name[i], mobile_number[i]]
#                 add_row_to_excel(filename, data)
#         else:
#             print("The lengths of person_name and mobile_number lists are not the same.")
#     else:
#         print("person_name and mobile_number should be lists.")



#     '''--------------------------------------------------------'''


def subImg_2_excel(filename):
    # for part1.jpg image
    image_1 = Image.open('left_side.jpg')
    # Extract text from the image
    text_1 = pytesseract.image_to_string(image_1)
    print(text_1)
    # Find Details from Extracted Text to Insert in Excel
    firm_name , firm_contact_number = extract_company_details(text_1)
    print("Firm Name : " ,firm_name)
    print("Firm Contact Number : " , firm_contact_number)

    # 'part1.jpg' extraction complete

    '''--------------------------------------------------------'''

    # for part2.jpg image
    image_2 = Image.open('right_side.jpg')
    # Extract text from the image
    text_2 = pytesseract.image_to_string(image_2)
    print(text_2)
    # Find Details from Extracted Text to Insert in Excel
    person_name , mobile_number = extract_name_and_phone(text_2)
    print(person_name)
    print(mobile_number)


    # Function to create a new Excel file with headers if it doesn't exist
    def create_excel_file(filename):
        workbook = Workbook()
        sheet = workbook.active
        headers = ["Firm Name", "Firm Contact Number", "Person Name", "Mobile Number"]
        sheet.append(headers)
        workbook.save(filename)

    # Function to add a new row to the existing Excel file
    def add_row_to_excel(filename, data):
        workbook = load_workbook(filename)
        sheet = workbook.active
        sheet.append(data)
        workbook.save(filename)

    # Check if the file exists, if not, create a new file with headers
    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        create_excel_file(filename)

    # Check if person_name and mobile_number are lists
    if isinstance(person_name, list) and isinstance(mobile_number, list):
        # Make sure the lengths of the lists are the same
        if len(person_name) == len(mobile_number):
            # Iterate over the lists and create a new row for each element
            for i in range(len(person_name)):
                data = [firm_name, firm_contact_number, person_name[i], mobile_number[i]]
                add_row_to_excel(filename, data)
        else:
            print("The lengths of person_name and mobile_number lists are not the same.")
    else:
        print("person_name and mobile_number should be lists.")



    '''--------------------------------------------------------'''